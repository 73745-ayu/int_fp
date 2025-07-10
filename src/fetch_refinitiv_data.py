# import refinitiv.data as rd
# import pandas as pd
# import numpy as np
# from datetime import datetime, timedelta
# from openpyxl import Workbook
# import uuid

# # Initialize Refinitiv session
# rd.open_session()

# # Configuration
# poa_input = "CY2026"
# poa_type = poa_input[:2]
# poa_year = int(poa_input[2:])
# companies = ['MCD']
# scale = 6
# pod_cutoff_estimate = pd.to_datetime("2025-06-01").date()
# today = pd.to_datetime("today").normalize().date()
# cutoff_date_POA = today + timedelta(days=3650)

# # Column names (dynamic)
# col_ticker = "Ticker"
# col_broker_name = "Broker Name"
# col_analyst_name = "Analyst Name"
# col_estimate_date = "Estimate Date"
# col_target_date = "Target Date"
# col_target_price = "Broker Target"
# col_dps = "DPS"
# col_div_yield = "Dividend Yield"
# col_div_yield_date = "Dividend Yield Date"
# col_ev = "EV"
# col_ebitda = "EBITDA"
# col_ebitda_margin = "EBITDA Margin"
# col_ebit_margin = "EBIT Margin"
# col_net_debt = "Net Debt"
# col_shares = "Shares Outstanding"
# col_rec_label = "Recommendation"
# col_rec_date = "Recommendation Date"
# col_revenue = "Revenue"
# col_rev_date = "Revenue Date"
# col_ebitda_date = "EBITDA Date"
# col_net_debt_date = "Net Debt Date"
# col_shares_date = "No. of Shares Outstanding Date"
# col_price = "Price"
# col_market_cap = "Market Cap"
# col_ev_ebitda = "EV/EBITDA"
# col_ebit = "EBIT"
# col_ebitda_12m_fwd = "EBITDA (12M Fwd)"

# # Broker overrides
# refinitiv_override = {
#     "PERMISSION DENIED 1342152": "SBI SECURITIES",
#     "PERMISSION DENIED 87408": "ROBERT W. BAIRD & CO",
#     "PERMISSION DENIED 937880": "TACHIBANA SECURITIES",
#     "PERMISSION DENIED 1120": "JEFFERIES",
#     "PERMISSION DENIED 1207112": "MELIUS RESEARCH",
#     "PERMISSION DENIED 1424952": "CFRA RESEARCH",
#     "PERMISSION DENIED 156648": "IWAICOSMO SECURITIES",
#     "PERMISSION DENIED 17472": "RBC CAPITAL MARKETS",
#     "PERMISSION DENIED 211744": "CROSS RESEARCH",
#     "PERMISSION DENIED 22760": "CLSA",
#     "PERMISSION DENIED 23440": "MIZUHO",
#     "PERMISSION DENIED 23816": "MORGAN STANLEY",
#     "PERMISSION DENIED 25632": "CITIGROUP",
#     "PERMISSION DENIED 266912": "KEPLER CHEUVREUX",
#     "PERMISSION DENIED 284328": "ARETE RESEARCH SERVICES LLP",
#     "PERMISSION DENIED 2880": "HSBC",
#     "PERMISSION DENIED 310016": "REDBURN ATLANTIC",
#     "PERMISSION DENIED 32": "BOFA",
#     "PERMISSION DENIED 32848": "BMO CAPITAL",
#     "PERMISSION DENIED 347360": "WOLFE RESEARCH",
#     "PERMISSION DENIED 36928": "JP MORGAN",
#     "PERMISSION DENIED 392": "DEUTSCHE BANK",
#     "PERMISSION DENIED 398136": "BARCLAYS",
#     "PERMISSION DENIED 483808": "HAITONG INTERNATIONAL",
#     "PERMISSION DENIED 495296": "MIZUHO",
#     "PERMISSION DENIED 512368": "SMBC NIKKO",
#     "PERMISSION DENIED 54992": "BERNSTEIN",
#     "PERMISSION DENIED 662336": "ALPHAVALUE",
#     "PERMISSION DENIED 696": "TD COWEN",
#     "PERMISSION DENIED 73704": "GOLDMAN SACHS",
#     "PERMISSION DENIED 7896": "DAIWA SECURITIES",
#     "PERMISSION DENIED 85152": "CANACCORD GENUITY"
# }

# def format_dates(df):
#     """Format date columns to a consistent string format"""
#     date_columns = [col for col in df.columns if "Date" in col]
#     for col in date_columns:
#         if col in df.columns:
#             df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d %b %y")
#     return df

# def consolidate_refinitiv_data(df, key_columns=None):
#     """Consolidate data by grouping on key columns and aggregating"""
#     if key_columns is None:
#         key_columns = [col_ticker, col_broker_name]
#         if col_estimate_date in df.columns:
#             key_columns.append(col_estimate_date)
    
#     # Check if key columns exist
#     existing_key_columns = [col for col in key_columns if col in df.columns]
#     if not existing_key_columns:
#         print(f"Warning: No key columns found in dataframe. Available columns: {list(df.columns)}")
#         return df
    
#     # Use only existing key columns
#     key_columns = existing_key_columns
    
#     # Identify numeric and non-numeric columns
#     numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
#     # Try to convert string columns to numeric if they look numeric
#     for col in df.columns:
#         if col not in numeric_cols and col not in key_columns:
#             try:
#                 temp = pd.to_numeric(df[col], errors='coerce')
#                 if temp.notna().mean() > 0.5:  # If more than 50% can be converted
#                     df[col] = temp
#                     numeric_cols.append(col)
#             except:
#                 pass
    
#     # Separate metadata columns
#     metadata_cols = [col for col in df.columns if col not in key_columns and col not in numeric_cols]
    
#     # Create aggregation dictionary
#     aggregations = {}
#     for col in numeric_cols:
#         if col not in key_columns:
#             aggregations[col] = lambda x: x.dropna().iloc[0] if not x.dropna().empty else np.nan
#     for col in metadata_cols:
#         aggregations[col] = 'first'
    
#     if not aggregations:
#         return df.drop_duplicates(subset=key_columns)
    
#     try:
#         result = df.groupby(key_columns, as_index=False).agg(aggregations)
#         return result
#     except Exception as e:
#         print(f"Error in consolidation: {e}")
#         return df.drop_duplicates(subset=key_columns)

# def apply_broker_overrides(df):
#     """Apply broker name overrides for permission denied entries"""
#     if col_broker_name in df.columns:
#         df[col_broker_name] = df[col_broker_name].astype(str).str.upper().str.strip()
#         for denied_key, broker_name in refinitiv_override.items():
#             df[col_broker_name] = df[col_broker_name].replace(denied_key, broker_name)
#     return df

# def get_metric_cy(metric_code, label, scale_on=True):
#     """Get calendar year metric data"""
#     try:
#         scale_str = f",Scale={scale}" if scale_on else ""
#         print(metric_code)
#         df = rd.get_data(
#             universe=companies,
#             fields=[f"{metric_code}.brokername", f"{metric_code}.date", f"{metric_code}{scale_str}"],
#             parameters={"Period": poa_input}
#         )
        
#         if df is None or df.empty:
#             print(f"Warning: No data returned for {label}")
#             return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])
        
#         df.columns = [col_ticker, col_broker_name, col_estimate_date, label]
        
#         df = apply_broker_overrides(df)
#         df[col_estimate_date] = pd.to_datetime(df[col_estimate_date], errors="coerce").dt.date
#         df = df[df[col_estimate_date] >= pod_cutoff_estimate]
        
#         return df.dropna(subset=[col_broker_name, col_estimate_date, label])
#     except Exception as e:
#         print(f"Error getting {label}: {e}")
#         return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])

# def get_metric_fy(metric_code, label, scale_on=True):
#     """Get fiscal year metric data"""
#     try:
#         scale_str = f",Scale={scale}" if scale_on else ""
#         df = rd.get_data(
#             universe=companies,
#             fields=[f"{metric_code}.brokername", f"{metric_code}.date", f"{metric_code}{scale_str}"],
#             parameters={"Period": poa_input}
#         )
        
#         if df is None or df.empty:
#             print(f"Warning: No data returned for {label}")
#             return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])
        
#         df.columns = [col_ticker, col_broker_name, col_estimate_date, label]
        
#         df = apply_broker_overrides(df)
#         df[col_estimate_date] = pd.to_datetime(df[col_estimate_date], errors="coerce").dt.date
#         df = df[df[col_estimate_date] >= pod_cutoff_estimate]
        
#         return df.dropna(subset=[col_broker_name, col_estimate_date, label])
#     except Exception as e:
#         print(f"Error getting {label}: {e}")
#         return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])

# def get_estimate_date(metric_date_field, label):
#     """Get estimate date for a metric"""
#     try:
#         df = rd.get_data(
#             universe=companies,
#             fields=[f"{metric_date_field}.brokername", f"{metric_date_field}.date"],
#             parameters={"Period": poa_input}
#         )
        
#         if df is None or df.empty:
#             return pd.DataFrame(columns=[col_ticker, col_broker_name, label])
        
#         df.columns = [col_ticker, col_broker_name, label]
        
#         df = apply_broker_overrides(df)
#         df[label] = pd.to_datetime(df[label], errors="coerce").dt.date
        
#         return df.dropna(subset=[col_broker_name, label])
#     except Exception as e:
#         print(f"Error getting estimate date for {label}: {e}")
#         return pd.DataFrame(columns=[col_ticker, col_broker_name, label])

# def create_multi_metric_forecast_summary(df, metrics, output_file="Multi_Metric_Forecast_Summary.xlsx"):
#     """Create Excel summary with forecast data and statistics"""
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Forecast Summary"
    
#     stats_measures = ["Median", "10th Percentile", "90th Percentile"]
#     current_row = 1
#     tickers = df[col_ticker].unique()
    
#     summary_dfs = {}
    
#     for ticker in tickers:
#         ticker_data = df[df[col_ticker] == ticker].copy()
        
#         # Write ticker header
#         ws.cell(row=current_row, column=1, value=f"{ticker} FORECAST PANEL")
#         current_row += 1
        
#         # Define display columns (only include existing ones)
#         display_cols = [
#             col_ticker, col_broker_name, col_analyst_name,
#             f"{col_revenue} {poa_input}", f"{col_ebitda} {poa_input}",
#             col_target_price, f"{col_net_debt} {poa_input}", f"{col_shares} {poa_input}",
#             f"{col_ebitda_margin} {poa_input}", f"{col_ev_ebitda} {poa_input}",
#             f"{poa_input} {col_div_yield}", col_ebitda_12m_fwd
#         ]
#         valid_cols = [col for col in display_cols if col in df.columns]
        
#         # Write headers
#         for col_idx, header in enumerate(valid_cols, 1):
#             ws.cell(row=current_row, column=col_idx, value=header)
#         current_row += 1
        
#         # Write data rows
#         for _, row in ticker_data.iterrows():
#             for col_idx, header in enumerate(valid_cols, 1):
#                 value = row.get(header, None)
#                 if pd.isna(value):
#                     value = None
#                 ws.cell(row=current_row, column=col_idx, value=value)
#             current_row += 1
        
#         current_row += 1
        
#         # Write summary statistics
#         ws.cell(row=current_row, column=1, value=f"Summary Statistics - {ticker}")
#         current_row += 1
        
#         ws.cell(row=current_row, column=1, value="Statistic")
#         valid_metrics = [m for m in metrics if m in df.columns]
#         for col_idx, metric in enumerate(valid_metrics, 2):
#             ws.cell(row=current_row, column=col_idx, value=metric)
#         current_row += 1
        
#         # Calculate and write statistics
#         summary_data = []
#         for stat in stats_measures:
#             ws.cell(row=current_row, column=1, value=stat)
#             stat_row = {"Statistic": stat}
            
#             for col_idx, metric in enumerate(valid_metrics, 2):
#                 values = ticker_data[metric].dropna()
#                 try:
#                     values = pd.to_numeric(values, errors='coerce').dropna()
#                     filtered_values = values[values != 0]
                    
#                     if len(filtered_values) == 0:
#                         value = None
#                     else:
#                         if stat == "Median":
#                             value = np.median(filtered_values)
#                         elif stat == "10th Percentile":
#                             value = np.percentile(filtered_values, 10)
#                         elif stat == "90th Percentile":
#                             value = np.percentile(filtered_values, 90)
#                         else:
#                             value = None
                    
#                     ws.cell(row=current_row, column=col_idx, value=value)
#                     stat_row[metric] = value
                    
#                     # Format percentage columns
#                     if value is not None and any(m in metric for m in ["Margin", "Dividend Yield"]):
#                         ws.cell(row=current_row, column=col_idx).number_format = '0.0%'
#                 except Exception as e:
#                     print(f"Error calculating {stat} for {metric}: {e}")
#                     ws.cell(row=current_row, column=col_idx, value=None)
#                     stat_row[metric] = None
            
#             summary_data.append(stat_row)
#             current_row += 1
        
#         # Store summary data
#         forecast_df = ticker_data[valid_cols] if valid_cols else ticker_data
#         summary_df = pd.DataFrame(summary_data).set_index("Statistic")
        
#         summary_dfs[ticker] = {
#             "Forecast Panel": forecast_df,
#             "Summary": summary_df
#         }
        
#         if ticker != tickers[-1]:
#             current_row += 2
    
#     # Save workbook
#     try:
#         wb.save(output_file)
#         print(f"Multi-metric forecast summary saved to {output_file}")
#     except Exception as e:
#         print(f"Error saving Excel file: {e}")
    
#     return summary_dfs

# def safe_divide(numerator, denominator):
#     """Safely divide two series, handling division by zero"""
#     try:
#         result = numerator / denominator
#         result = result.replace([np.inf, -np.inf], np.nan)
#         return result
#     except:
#         return pd.Series([np.nan] * len(numerator))

# # Main data processing
# print("Starting data retrieval...")

# # Define metrics to retrieve
# metrics = {
#     f"{col_revenue} {poa_input}": "TR.RevenueEstValue",
#     f"{col_ebitda} {poa_input}": "TR.EBITDAEstValue",
#     f"{col_ebit} {poa_input}": "TR.EBITEstValue",
#     f"{col_net_debt} {poa_input}": "TR.NetDebtEstValue",
#     f"{col_dps} {poa_input}": "TR.DPSEstValue",
# }

# # Retrieve metric data
# raw_data_frames = {}
# data_frames = []

# for label, code in metrics.items():
#     print(f"Retrieving {label}...")
#     if poa_type == "CY":
#         df = get_metric_cy(code, label, scale_on=False if col_dps in label else True)
#     else:
#         df = get_metric_fy(code, label, scale_on=False if col_dps in label else True)
    
#     if not df.empty:
#         data_frames.append(df)
#         raw_data_frames[label] = df.copy()
#         print(f"Retrieved {len(df)} records for {label}")
#     else:
#         print(f"No data retrieved for {label}")

# # Get shares data
# print("Retrieving shares data...")
# try:
#     shares_df = rd.get_data(
#         universe=companies,
#         fields=[f"TR.NumberOfSharesOutstanding.brokername", f"TR.NumberOfSharesOutstanding(Period={poa_input})"]
#     )
    
#     if shares_df is not None and not shares_df.empty:
#         shares_df.columns = [col_ticker, col_broker_name, f"{col_shares} {poa_input}"]
#         # Convert shares to millions
#         shares_df[f"{col_shares} {poa_input}"] = pd.to_numeric(shares_df[f"{col_shares} {poa_input}"], errors='coerce') / 1_000_000
#         shares_df = apply_broker_overrides(shares_df)
#         shares_df = shares_df.dropna(subset=[col_broker_name, f"{col_shares} {poa_input}"]).drop_duplicates(subset=[col_ticker, col_broker_name])
#         data_frames.append(shares_df)
#         raw_data_frames[f"{col_shares} {poa_input}"] = shares_df.copy()
#         print(f"Retrieved {len(shares_df)} records for shares data")
# except Exception as e:
#     print(f"Error retrieving shares data: {e}")

# # Get target price data
# print("Retrieving target price data...")
# try:
#     tp_df = rd.get_data(
#         universe=companies,
#         fields=["TR.PriceTargetMedian"],
#         parameters={"Period": poa_input}
#     )
    
#     if tp_df is not None and not tp_df.empty:
#         tp_df.columns = [col_ticker, col_target_price]
#         # Convert from pence to GBP if needed
#         tp_df[col_target_price] = pd.to_numeric(tp_df[col_target_price], errors='coerce') / 100
#         data_frames.append(tp_df)
#         raw_data_frames[col_target_price] = tp_df.copy()
#         print(f"Retrieved {len(tp_df)} records for target price data")
# except Exception as e:
#     print(f"Error retrieving target price data: {e}")

# # Get recommendation data
# print("Retrieving recommendation data...")
# try:
#     rec_df = rd.get_data(
#         universe=companies,
#         fields=["TR.BrkRecEstBrokerName", "TR.BrkRecLabel", "TR.BrkRecLabelEstDate"],
#         parameters={"Period": poa_input}
#     )
    
#     if rec_df is not None and not rec_df.empty:
#         rec_df.columns = [col_ticker, col_broker_name, col_rec_label, col_rec_date]
#         rec_df = apply_broker_overrides(rec_df)
#         rec_df[col_rec_date] = pd.to_datetime(rec_df[col_rec_date], errors="coerce").dt.date
#         rec_df = rec_df.drop_duplicates(subset=[col_ticker, col_broker_name])
#         data_frames.append(rec_df)
#         raw_data_frames[col_rec_label] = rec_df.copy()
#         print(f"Retrieved {len(rec_df)} records for recommendation data")
# except Exception as e:
#     print(f"Error retrieving recommendation data: {e}")

# # Get date fields
# print("Retrieving date fields...")
# date_fields = [
#     ("TR.RevenueEstDate", f"{poa_input} {col_rev_date}"),
#     ("TR.EBITDAEstDate", f"{poa_input} {col_ebitda_date}"),
#     ("TR.NetDebtEstDate", f"{poa_input} {col_net_debt_date}"),
#     ("TR.NumberOfSharesOutstanding", f"{poa_input} {col_shares_date}")
# ]

# for field, label in date_fields:
#     date_df = get_estimate_date(field, label)
#     if not date_df.empty:
#         data_frames.append(date_df)

# # Consolidate all data
# print("Consolidating data...")
# if not data_frames:
#     print("No data frames to consolidate")
#     exit()

# # Create base panel from all ticker-broker combinations
# all_tickers_brokers = pd.DataFrame()
# for df in data_frames:
#     if col_broker_name in df.columns and col_ticker in df.columns:
#         temp_df = df[[col_ticker, col_broker_name]].drop_duplicates()
#         all_tickers_brokers = pd.concat([all_tickers_brokers, temp_df], ignore_index=True)

# if all_tickers_brokers.empty:
#     print("No ticker-broker combinations found")
#     exit()

# all_tickers_brokers = all_tickers_brokers.drop_duplicates()
# print(f"Found {len(all_tickers_brokers)} unique ticker-broker combinations")

# # Merge all data frames
# panel = all_tickers_brokers.copy()
# for df in data_frames:
#     if df.empty:
#         continue
        
#     if col_broker_name not in df.columns:
#         # Merge on ticker only
#         if col_ticker in df.columns:
#             panel = pd.merge(panel, df, on=col_ticker, how="left")
#     else:
#         # Merge on ticker and broker
#         common_cols = [col for col in [col_ticker, col_broker_name] if col in df.columns]
#         if common_cols:
#             panel = pd.merge(panel, df, on=common_cols, how="left", suffixes=('', '_drop'))

# # Remove duplicate columns
# panel = panel[[col for col in panel.columns if not col.endswith('_drop')]]

# # Consolidate the panel
# key_columns = [col_ticker, col_broker_name]
# if col_estimate_date in panel.columns:
#     key_columns.append(col_estimate_date)

# panel = consolidate_refinitiv_data(panel, key_columns=key_columns)

# print(f"Panel shape after consolidation: {panel.shape}")
# print(f"Panel columns: {list(panel.columns)}")

# # Calculate derived metrics
# print("Calculating derived metrics...")

# # Market Cap = Shares * Target Price
# if all(col in panel.columns for col in [f"{col_shares} {poa_input}", col_target_price]):
#     panel[f"{col_market_cap} {poa_input}"] = (
#         pd.to_numeric(panel[f"{col_shares} {poa_input}"], errors='coerce') * 
#         pd.to_numeric(panel[col_target_price], errors='coerce')
#     )
#     print("Calculated Market Cap")

# # EV = Market Cap + Net Debt
# if all(col in panel.columns for col in [f"{col_market_cap} {poa_input}", f"{col_net_debt} {poa_input}"]):
#     panel[f"{col_ev} {poa_input}"] = (
#         pd.to_numeric(panel[f"{col_market_cap} {poa_input}"], errors='coerce') + 
#         pd.to_numeric(panel[f"{col_net_debt} {poa_input}"], errors='coerce')
#     )
#     print("Calculated EV")

# # EV/EBITDA = EV / EBITDA
# if all(col in panel.columns for col in [f"{col_ev} {poa_input}", f"{col_ebitda} {poa_input}"]):
#     panel[f"{col_ev_ebitda} {poa_input}"] = safe_divide(
#         pd.to_numeric(panel[f"{col_ev} {poa_input}"], errors='coerce'),
#         pd.to_numeric(panel[f"{col_ebitda} {poa_input}"], errors='coerce')
#     )
#     print("Calculated EV/EBITDA")

# # EBITDA Margin = EBITDA / Revenue
# if all(col in panel.columns for col in [f"{col_ebitda} {poa_input}", f"{col_revenue} {poa_input}"]):
#     panel[f"{col_ebitda_margin} {poa_input}"] = safe_divide(
#         pd.to_numeric(panel[f"{col_ebitda} {poa_input}"], errors='coerce'),
#         pd.to_numeric(panel[f"{col_revenue} {poa_input}"], errors='coerce')
#     )
#     print("Calculated EBITDA Margin")

# # Dividend Yield = DPS / Target Price
# if all(col in panel.columns for col in [f"{col_dps} {poa_input}", col_target_price]):
#     panel[f"{poa_input} {col_div_yield}"] = safe_divide(
#         pd.to_numeric(panel[f"{col_dps} {poa_input}"], errors='coerce'),
#         pd.to_numeric(panel[col_target_price], errors='coerce')
#     )
#     panel[f"{poa_input} {col_div_yield_date}"] = today
#     print("Calculated Dividend Yield")

# # EBITDA 12M Forward (same as EBITDA estimate)
# if f"{col_ebitda} {poa_input}" in panel.columns:
#     panel[col_ebitda_12m_fwd] = panel[f"{col_ebitda} {poa_input}"]
#     print("Set EBITDA 12M Forward")

# # Convert numerical columns
# numerical_cols = [
#     f"{col_revenue} {poa_input}",
#     f"{col_ebitda} {poa_input}",
#     col_target_price,
#     f"{col_net_debt} {poa_input}",
#     f"{col_shares} {poa_input}",
#     f"{col_ebitda_margin} {poa_input}",
#     f"{col_ev_ebitda} {poa_input}",
#     f"{poa_input} {col_div_yield}",
#     col_ebitda_12m_fwd,
#     f"{col_market_cap} {poa_input}",
#     f"{col_ev} {poa_input}"
# ]

# for col in numerical_cols:
#     if col in panel.columns:
#         panel[col] = pd.to_numeric(panel[col], errors='coerce')

# # Remove rows with all NaN values in key metrics
# key_metrics = [col for col in numerical_cols if col in panel.columns]
# if key_metrics:
#     panel = panel.dropna(subset=key_metrics, how='all')

# print(f"Final panel shape: {panel.shape}")

# # Display sample data for validation
# if not panel.empty:
#     print("\nSample data from final panel:")
#     display_cols = [col for col in [
#         col_ticker, col_broker_name, 
#         f"{col_shares} {poa_input}", col_target_price, 
#         f"{col_market_cap} {poa_input}", f"{col_net_debt} {poa_input}", 
#         f"{col_ev} {poa_input}", f"{col_ebitda} {poa_input}", 
#         f"{col_ev_ebitda} {poa_input}"
#     ] if col in panel.columns]
    
#     print(panel[display_cols].head())

# # Format dates
# panel = format_dates(panel)

# # Define metrics to analyze
# metrics_to_analyze = [col for col in numerical_cols if col in panel.columns]

# if panel.empty:
#     print("No data available for analysis")
# else:
#     # Create summary
#     print("Creating forecast summary...")
#     summary_dfs = create_multi_metric_forecast_summary(
#         panel, 
#         metrics_to_analyze, 
#         output_file="Combined_Forecast_Summary_With_Linking.xlsx"
#     )
    
#     # Display results
#     for ticker, dfs in summary_dfs.items():
#         print(f"\n{'='*50}")
#         print(f"RESULTS FOR {ticker}")
#         print(f"{'='*50}")
        
#         print(f"\nForecast Panel Shape: {dfs['Forecast Panel'].shape}")
#         print(f"Available Metrics: {list(dfs['Forecast Panel'].columns)}")
        
#         # Display key statistics
#         key_stats_metrics = [
#             f"{col_revenue} {poa_input}", 
#             f"{col_ebitda_margin} {poa_input}", 
#             f"{col_ev_ebitda} {poa_input}"
#         ]
        
#         for metric in key_stats_metrics:
#             if metric in dfs["Forecast Panel"].columns:
#                 values = dfs["Forecast Panel"][metric].dropna()
#                 if not values.empty:
#                     try:
#                         values = pd.to_numeric(values, errors='coerce').dropna()
#                         if not values.empty:
#                             p10 = np.percentile(values, 10)
#                             median = np.median(values)
#                             p90 = np.percentile(values, 90)
                            
#                             is_percentage = 'Margin' in metric or 'Dividend Yield' in metric
#                             suffix = '%' if is_percentage else ''
                            
#                             print(f"\n{metric}:")
#                             print(f"  10th percentile: {p10:.2f}{suffix}")
#                             print(f"  Median: {median:.2f}{suffix}")
#                             print(f"  90th percentile: {p90:.2f}{suffix}")
#                     except Exception as e:
#                         print(f"Error calculating statistics for {metric}: {e}")
        
#         print(f"\nSummary Statistics Table:")
#         if not dfs["Summary"].empty:
#             print(dfs["Summary"])
#         else:
#             print("No summary statistics available")

# print("\nScript execution completed successfully!")

# # Close Refinitiv session
# try:
#     rd.close_session()
#     print("Refinitiv session closed.")
# except Exception as e:
#     print(f"Error closing Refinitiv session: {e}")
import refinitiv.data as rd
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
import uuid

# Initialize Refinitiv session
rd.open_session()

# Configuration
poa_input = "CY2026"
poa_type = poa_input[:2]
poa_year = int(poa_input[2:])
companies = ['CRDA.L']
scale = 6  # For millions
pod_cutoff_estimate = pd.to_datetime("2025-06-01").date()
today = pd.to_datetime("today").normalize().date()
cutoff_date_POA = today + timedelta(days=3650)

# Column names
col_ticker = "Ticker"
col_broker_name = "Broker Name"
col_analyst_name = "Analyst Name"
col_estimate_date = "Estimate Date"
col_target_date = "Target Date"
col_target_price = "Broker Target"
col_dps = "DPS"
col_div_yield = "Dividend Yield"
col_div_yield_date = "Dividend Yield Date"
col_ev = "EV"
col_ebitda = "EBITDA"
col_ebitda_margin = "EBITDA Margin"
col_ebit_margin = "EBIT Margin"
col_net_debt = "Net Debt"
col_shares = "Shares Outstanding"
col_rec_label = "Recommendation"
col_rec_date = "Recommendation Date"
col_revenue = "Revenue"
col_rev_date = "Revenue Date"
col_ebitda_date = "EBITDA Date"
col_net_debt_date = "Net Debt Date"
col_shares_date = "No. of Shares Outstanding Date"
col_price = "Price"
col_market_cap = "Market Cap"
col_ev_ebitda = "EV/EBITDA"
col_ebit = "EBIT"
col_ebitda_12m_fwd = "EBITDA (12M Fwd)"

# Broker overrides
refinitiv_override = {
    "PERMISSION DENIED 1342152": "SBI SECURITIES",
    "PERMISSION DENIED 87408": "ROBERT W. BAIRD & CO",
    "PERMISSION DENIED 937880": "TACHIBANA SECURITIES",
    "PERMISSION DENIED 1120": "JEFFERIES",
    "PERMISSION DENIED 1207112": "MELIUS RESEARCH",
    "PERMISSION DENIED 1424952": "CFRA RESEARCH",
    "PERMISSION DENIED 156648": "IWAICOSMO SECURITIES",
    "PERMISSION DENIED 17472": "RBC CAPITAL MARKETS",
    "PERMISSION DENIED 211744": "CROSS RESEARCH",
    "PERMISSION DENIED 22760": "CLSA",
    "PERMISSION DENIED 23440": "MIZUHO",
    "PERMISSION DENIED 23816": "MORGAN STANLEY",
    "PERMISSION DENIED 25632": "CITIGROUP",
    "PERMISSION DENIED 266912": "KEPLER CHEUVREUX",
    "PERMISSION DENIED 284328": "ARETE RESEARCH SERVICES LLP",
    "PERMISSION DENIED 2880": "HSBC",
    "PERMISSION DENIED 310016": "REDBURN ATLANTIC",
    "PERMISSION DENIED 32": "BOFA",
    "PERMISSION DENIED 32848": "BMO CAPITAL",
    "PERMISSION DENIED 347360": "WOLFE RESEARCH",
    "PERMISSION DENIED 36928": "JP MORGAN",
    "PERMISSION DENIED 392": "DEUTSCHE BANK",
    "PERMISSION DENIED 398136": "BARCLAYS",
    "PERMISSION DENIED 483808": "HAITONG INTERNATIONAL",
    "PERMISSION DENIED 495296": "MIZUHO",
    "PERMISSION DENIED 512368": "SMBC NIKKO",
    "PERMISSION DENIED 54992": "BERNSTEIN",
    "PERMISSION DENIED 662336": "ALPHAVALUE",
    "PERMISSION DENIED 696": "TD COWEN",
    "PERMISSION DENIED 73704": "GOLDMAN SACHS",
    "PERMISSION DENIED 7896": "DAIWA SECURITIES",
    "PERMISSION DENIED 85152": "CANACCORD GENUITY"
}

def format_dates(df):
    """Format date columns to a consistent string format"""
    date_columns = [col for col in df.columns if "Date" in col]
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d %b %y")
    return df

def consolidate_refinitiv_data(df, key_columns=None):
    """Consolidate data by grouping on key columns and aggregating"""
    if key_columns is None:
        key_columns = [col_ticker, col_broker_name]
        if col_estimate_date in df.columns:
            key_columns.append(col_estimate_date)
    
    existing_key_columns = [col for col in key_columns if col in df.columns]
    if not existing_key_columns:
        print(f"Warning: No key columns found in dataframe. Available columns: {list(df.columns)}")
        return df
    
    key_columns = existing_key_columns
    
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    for col in df.columns:
        if col not in numeric_cols and col not in key_columns:
            try:
                temp = pd.to_numeric(df[col], errors='coerce')
                if temp.notna().mean() > 0.5:
                    df[col] = temp
                    numeric_cols.append(col)
            except:
                pass
    
    metadata_cols = [col for col in df.columns if col not in key_columns and col not in numeric_cols]
    
    aggregations = {}
    for col in numeric_cols:
        if col not in key_columns:
            aggregations[col] = lambda x: x.dropna().iloc[0] if not x.dropna().empty else np.nan
    for col in metadata_cols:
        aggregations[col] = 'first'
    
    if not aggregations:
        return df.drop_duplicates(subset=key_columns)
    
    try:
        result = df.groupby(key_columns, as_index=False).agg(aggregations)
        return result
    except Exception as e:
        print(f"Error in consolidation: {e}")
        return df.drop_duplicates(subset=key_columns)

def apply_broker_overrides(df):
    """Apply broker name overrides for permission denied entries"""
    if col_broker_name in df.columns:
        df[col_broker_name] = df[col_broker_name].astype(str).str.upper().str.strip()
        for denied_key, broker_name in refinitiv_override.items():
            df[col_broker_name] = df[col_broker_name].replace(denied_key, broker_name)
    return df

def get_metric_cy(metric_code, label, scale_on=True):
    """Get calendar year metric data"""
    try:
        scale_str = f",Scale={scale}" if scale_on else ""
        print(f"Retrieving {metric_code} for {label}")
        df = rd.get_data(
            universe=companies,
            fields=[f"{metric_code}.brokername", f"{metric_code}.date", f"{metric_code}{scale_str}"],
            parameters={"Period": poa_input}
        )
        
        if df is None or df.empty:
            print(f"Warning: No data returned for {label}")
            return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])
        
        df.columns = [col_ticker, col_broker_name, col_estimate_date, label]
        
        df = apply_broker_overrides(df)
        df[col_estimate_date] = pd.to_datetime(df[col_estimate_date], errors="coerce").dt.date
        df = df[df[col_estimate_date] >= pod_cutoff_estimate]
        
        return df.dropna(subset=[col_broker_name, col_estimate_date, label])
    except Exception as e:
        print(f"Error getting {label}: {e}")
        return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])

def get_metric_fy(metric_code, label, scale_on=True):
    """Get fiscal year metric data"""
    try:
        scale_str = f",Scale={scale}" if scale_on else ""
        df = rd.get_data(
            universe=companies,
            fields=[f"{metric_code}.brokername", f"{metric_code}.date", f"{metric_code}{scale_str}"],
            parameters={"Period": poa_input}
        )
        
        if df is None or df.empty:
            print(f"Warning: No data returned for {label}")
            return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])
        
        df.columns = [col_ticker, col_broker_name, col_estimate_date, label]
        
        df = apply_broker_overrides(df)
        df[col_estimate_date] = pd.to_datetime(df[col_estimate_date], errors="coerce").dt.date
        df = df[df[col_estimate_date] >= pod_cutoff_estimate]
        
        return df.dropna(subset=[col_broker_name, col_estimate_date, label])
    except Exception as e:
        print(f"Error getting {label}: {e}")
        return pd.DataFrame(columns=[col_ticker, col_broker_name, col_estimate_date, label])

def get_estimate_date(metric_date_field, label):
    """Get estimate date for a metric"""
    try:
        df = rd.get_data(
            universe=companies,
            fields=[f"{metric_date_field}.brokername", f"{metric_date_field}.date"],
            parameters={"Period": poa_input}
        )
        
        if df is None or df.empty:
            return pd.DataFrame(columns=[col_ticker, col_broker_name, label])
        
        df.columns = [col_ticker, col_broker_name, label]
        
        df = apply_broker_overrides(df)
        df[label] = pd.to_datetime(df[label], errors="coerce").dt.date
        
        return df.dropna(subset=[col_broker_name, label])
    except Exception as e:
        print(f"Error getting estimate date for {label}: {e}")
        return pd.DataFrame(columns=[col_ticker, col_broker_name, label])

def create_multi_metric_forecast_summary(df, metrics, output_file="Multi_Metric_Forecast_Summary.xlsx"):
    """Create Excel summary with forecast data and statistics"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Summary"
    
    stats_measures = ["Median", "10th Percentile", "90th Percentile"]
    current_row = 1
    tickers = df[col_ticker].unique()
    
    summary_dfs = {}
    
    for ticker in tickers:
        ticker_data = df[df[col_ticker] == ticker].copy()
        
        ws.cell(row=current_row, column=1, value=f"{ticker} FORECAST PANEL")
        current_row += 1
        
        display_cols = [
            col_ticker, col_broker_name, col_analyst_name,
            f"{col_revenue} {poa_input}", f"{col_ebitda} {poa_input}",
            col_target_price, f"{col_net_debt} {poa_input}", f"{col_shares} {poa_input}",
            f"{col_ebitda_margin} {poa_input}", f"{col_ev_ebitda} {poa_input}",
            f"{poa_input} {col_div_yield}", col_ebitda_12m_fwd,
            f"{col_market_cap} {poa_input}", f"{col_ev} {poa_input}"
        ]
        valid_cols = [col for col in display_cols if col in df.columns]
        
        for col_idx, header in enumerate(valid_cols, 1):
            ws.cell(row=current_row, column=col_idx, value=header)
        current_row += 1
        
        for _, row in ticker_data.iterrows():
            for col_idx, header in enumerate(valid_cols, 1):
                value = row.get(header, None)
                if pd.isna(value):
                    value = None
                elif isinstance(value, (int, float)) and not pd.isna(value):
                    if any(keyword in header for keyword in ["Margin", "Dividend Yield"]):
                        ws.cell(row=current_row, column=col_idx, value=value)
                        ws.cell(row=current_row, column=col_idx).number_format = '0.0%'
                    else:
                        ws.cell(row=current_row, column=col_idx, value=value)
                else:
                    ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Summary Statistics - {ticker}")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Statistic")
        valid_metrics = [m for m in metrics if m in df.columns]
        for col_idx, metric in enumerate(valid_metrics, 2):
            ws.cell(row=current_row, column=col_idx, value=metric)
        current_row += 1
        
        summary_data = []
        for stat in stats_measures:
            ws.cell(row=current_row, column=1, value=stat)
            stat_row = {"Statistic": stat}
            
            for col_idx, metric in enumerate(valid_metrics, 2):
                values = ticker_data[metric].dropna()
                try:
                    values = pd.to_numeric(values, errors='coerce').dropna()
                    filtered_values = values[values != 0]
                    
                    if len(filtered_values) == 0:
                        value = None
                    else:
                        if stat == "Median":
                            value = np.median(filtered_values)
                        elif stat == "10th Percentile":
                            value = np.percentile(filtered_values, 10)
                        elif stat == "90th Percentile":
                            value = np.percentile(filtered_values, 90)
                        else:
                            value = None
                    
                    if value is not None:
                        ws.cell(row=current_row, column=col_idx, value=value)
                        if any(keyword in metric for keyword in ["Margin", "Dividend Yield"]):
                            ws.cell(row=current_row, column=col_idx).number_format = '0.0%'
                    else:
                        ws.cell(row=current_row, column=col_idx, value=None)
                    
                    stat_row[metric] = value
                    
                except Exception as e:
                    print(f"Error calculating {stat} for {metric}: {e}")
                    ws.cell(row=current_row, column=col_idx, value=None)
                    stat_row[metric] = None
            
            summary_data.append(stat_row)
            current_row += 1
        
        forecast_df = ticker_data[valid_cols] if valid_cols else ticker_data
        summary_df = pd.DataFrame(summary_data).set_index("Statistic")
        
        summary_dfs[ticker] = {
            "Forecast Panel": forecast_df,
            "Summary": summary_df
        }
        
        if ticker != tickers[-1]:
            current_row += 2
    
    try:
        wb.save(output_file)
        print(f"Multi-metric forecast summary saved to {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
    
    return summary_dfs

def safe_divide(numerator, denominator):
    """Safely divide two series, handling division by zero"""
    try:
        num = pd.to_numeric(numerator, errors='coerce')
        den = pd.to_numeric(denominator, errors='coerce')
        den = den.replace(0, np.nan)
        result = num / den
        result = result.replace([np.inf, -np.inf], np.nan)
        return result
    except Exception as e:
        print(f"Error in safe_divide: {e}")
        return pd.Series([np.nan] * len(numerator) if hasattr(numerator, '__len__') else np.nan)

# Main data processing
print("Starting data retrieval...")

metrics = {
    f"{col_revenue} {poa_input}": "TR.RevenueEstValue",
    f"{col_ebitda} {poa_input}": "TR.EBITDAEstValue",
    f"{col_ebit} {poa_input}": "TR.EBITEstValue",
    f"{col_net_debt} {poa_input}": "TR.NetDebtEstValue",
    f"{col_dps} {poa_input}": "TR.DPSEstValue",
}

raw_data_frames = {}
data_frames = []

for label, code in metrics.items():
    print(f"Retrieving {label}...")
    if poa_type == "CY":
        df = get_metric_cy(code, label, scale_on=False if col_dps in label else True)
    else:
        df = get_metric_fy(code, label, scale_on=False if col_dps in label else True)
    
    if not df.empty:
        data_frames.append(df)
        raw_data_frames[label] = df.copy()
        print(f"Retrieved {len(df)} records for {label}")
    else:
        print(f"No data retrieved for {label}")

print("Retrieving shares data...")
try:
    shares_df = rd.get_data(
        universe=companies,
        fields=[f"TR.NumberOfSharesOutstanding.brokername", f"TR.NumberOfSharesOutstanding.date", f"TR.NumberOfSharesOutstanding,Scale={scale}"],
        parameters={"Period": poa_input}
    )
    
    if shares_df is not None and not shares_df.empty:
        shares_df.columns = [col_ticker, col_broker_name, col_estimate_date, f"{col_shares} {poa_input}"]
        shares_df = apply_broker_overrides(shares_df)
        shares_df[col_estimate_date] = pd.to_datetime(shares_df[col_estimate_date], errors="coerce").dt.date
        shares_df = shares_df[shares_df[col_estimate_date] >= pod_cutoff_estimate]
        shares_df = shares_df.dropna(subset=[col_broker_name, f"{col_shares} {poa_input}"])
        data_frames.append(shares_df)
        raw_data_frames[f"{col_shares} {poa_input}"] = shares_df.copy()
        print(f"Retrieved {len(shares_df)} records for shares data")
except Exception as e:
    print(f"Error retrieving shares data: {e}")

print("Retrieving target price data...")
try:
    tp_df = rd.get_data(
        universe=companies,
        fields=["TR.PriceTargetValue.brokername", "TR.PriceTargetValue.date", "TR.PriceTargetValue"],
        parameters={"Period": poa_input}
    )
    
    if tp_df is not None and not tp_df.empty:
        tp_df.columns = [col_ticker, col_broker_name, col_estimate_date, col_target_price]
        tp_values = pd.to_numeric(tp_df[col_target_price], errors='coerce')
        if tp_values.median() > 100:
            tp_df[col_target_price] = tp_values / 100
        else:
            tp_df[col_target_price] = tp_values
            
        tp_df = apply_broker_overrides(tp_df)
        tp_df[col_estimate_date] = pd.to_datetime(tp_df[col_estimate_date], errors="coerce").dt.date
        tp_df = tp_df[tp_df[col_estimate_date] >= pod_cutoff_estimate]
        tp_df = tp_df.dropna(subset=[col_broker_name, col_target_price])
        data_frames.append(tp_df)
        raw_data_frames[col_target_price] = tp_df.copy()
        print(f"Retrieved {len(tp_df)} records for target price data")
except Exception as e:
    print(f"Error retrieving target price data: {e}")

print("Retrieving recommendation data...")
try:
    rec_df = rd.get_data(
        universe=companies,
        fields=["TR.BrkRecEstBrokerName", "TR.BrkRecLabel", "TR.BrkRecLabelEstDate"],
        parameters={"Period": poa_input}
    )
    
    if rec_df is not None and not rec_df.empty:
        rec_df.columns = [col_ticker, col_broker_name, col_rec_label, col_rec_date]
        rec_df = apply_broker_overrides(rec_df)
        rec_df[col_rec_date] = pd.to_datetime(rec_df[col_rec_date], errors="coerce").dt.date
        rec_df = rec_df.drop_duplicates(subset=[col_ticker, col_broker_name])
        data_frames.append(rec_df)
        raw_data_frames[col_rec_label] = rec_df.copy()
        print(f"Retrieved {len(rec_df)} records for recommendation data")
except Exception as e:
    print(f"Error retrieving recommendation data: {e}")

print("Retrieving date fields...")
date_fields = [
    ("TR.RevenueEstDate", f"{poa_input} {col_rev_date}"),
    ("TR.EBITDAEstDate", f"{poa_input} {col_ebitda_date}"),
    ("TR.NetDebtEstDate", f"{poa_input} {col_net_debt_date}"),
    ("TR.NumberOfSharesOutstanding", f"{poa_input} {col_shares_date}")
]

for field, label in date_fields:
    date_df = get_estimate_date(field, label)
    if not date_df.empty:
        data_frames.append(date_df)

print("Consolidating data...")
if not data_frames:
    print("No data frames to consolidate")
    exit()

all_tickers_brokers = pd.DataFrame()
for df in data_frames:
    if col_broker_name in df.columns and col_ticker in df.columns:
        temp_df = df[[col_ticker, col_broker_name]].drop_duplicates()
        all_tickers_brokers = pd.concat([all_tickers_brokers, temp_df], ignore_index=True)

if all_tickers_brokers.empty:
    print("No ticker-broker combinations found")
    exit()

all_tickers_brokers = all_tickers_brokers.drop_duplicates()
print(f"Found {len(all_tickers_brokers)} unique ticker-broker combinations")

panel = all_tickers_brokers.copy()
for df in data_frames:
    if df.empty:
        continue
        
    if col_broker_name not in df.columns:
        if col_ticker in df.columns:
            panel = pd.merge(panel, df, on=col_ticker, how="left")
    else:
        common_cols = [col for col in [col_ticker, col_broker_name] if col in df.columns]
        if common_cols:
            panel = pd.merge(panel, df, on=common_cols, how="left", suffixes=('', '_drop'))

panel = panel[[col for col in panel.columns if not col.endswith('_drop')]]

key_columns = [col_ticker, col_broker_name]
if col_estimate_date in panel.columns:
    key_columns.append(col_estimate_date)

panel = consolidate_refinitiv_data(panel, key_columns=key_columns)

print(f"Panel shape after consolidation: {panel.shape}")
print(f"Panel columns: {list(panel.columns)}")

print("Calculating derived metrics...")

# Market Cap = Shares (millions) * Target Price
if all(col in panel.columns for col in [f"{col_shares} {poa_input}", col_target_price]):
    shares_numeric = pd.to_numeric(panel[f"{col_shares} {poa_input}"], errors='coerce')
    price_numeric = pd.to_numeric(panel[col_target_price], errors='coerce')
    panel[f"{col_market_cap} {poa_input}"] = shares_numeric * price_numeric
    print(f"Calculated Market Cap - Sample values: {panel[f'{col_market_cap} {poa_input}'].dropna().head().values}")

# EV = Market Cap + Net Debt
if all(col in panel.columns for col in [f"{col_market_cap} {poa_input}", f"{col_net_debt} {poa_input}"]):
    mkt_cap_numeric = pd.to_numeric(panel[f"{col_market_cap} {poa_input}"], errors='coerce')
    net_debt_numeric = pd.to_numeric(panel[f"{col_net_debt} {poa_input}"], errors='coerce')
    panel[f"{col_ev} {poa_input}"] = mkt_cap_numeric + net_debt_numeric
    print(f"Calculated EV - Sample values: {panel[f'{col_ev} {poa_input}'].dropna().head().values}")

# EV/EBITDA
if all(col in panel.columns for col in [f"{col_ev} {poa_input}", f"{col_ebitda} {poa_input}"]):
    panel[f"{col_ev_ebitda} {poa_input}"] = safe_divide(
        panel[f"{col_ev} {poa_input}"],
        panel[f"{col_ebitda} {poa_input}"]
    )
    print(f"Calculated EV/EBITDA - Sample values: {panel[f'{col_ev_ebitda} {poa_input}'].dropna().head().values}")

# EBITDA Margin = EBITDA / Revenue
if all(col in panel.columns for col in [f"{col_ebitda} {poa_input}", f"{col_revenue} {poa_input}"]):
    ebitda_numeric = pd.to_numeric(panel[f"{col_ebitda} {poa_input}"], errors='coerce')
    revenue_numeric = pd.to_numeric(panel[f"{col_revenue} {poa_input}"], errors='coerce')
    panel[f"{col_ebitda_margin} {poa_input}"] = safe_divide(ebitda_numeric, revenue_numeric)
    print(f"Calculated EBITDA Margin - Sample values: {panel[f'{col_ebitda_margin} {poa_input}'].dropna().head().values}")

# Dividend Yield = DPS / Target Price
if all(col in panel.columns for col in [f"{col_dps} {poa_input}", col_target_price]):
    dps_numeric = pd.to_numeric(panel[f"{col_dps} {poa_input}"], errors='coerce')
    price_numeric = pd.to_numeric(panel[col_target_price], errors='coerce')
    panel[f"{poa_input} {col_div_yield}"] = safe_divide(dps_numeric, price_numeric)
    panel[f"{poa_input} {col_div_yield_date}"] = today
    print(f"Calculated Dividend Yield - Sample values: {panel[f'{poa_input} {col_div_yield}'].dropna().head().values}")

# EBITDA 12M Forward
if f"{col_ebitda} {poa_input}" in panel.columns:
    panel[col_ebitda_12m_fwd] = panel[f"{col_ebitda} {poa_input}"]
    print("Set EBITDA 12M Forward")

numerical_cols = [
    f"{col_revenue} {poa_input}",
    f"{col_ebitda} {poa_input}",
    col_target_price,
    f"{col_net_debt} {poa_input}",
    f"{col_shares} {poa_input}",
    f"{col_ebitda_margin} {poa_input}",
    f"{col_ev_ebitda} {poa_input}",
    f"{poa_input} {col_div_yield}",
    col_ebitda_12m_fwd,
    f"{col_market_cap} {poa_input}",
    f"{col_ev} {poa_input}"
]

for col in numerical_cols:
    if col in panel.columns:
        panel[col] = pd.to_numeric(panel[col], errors='coerce')

key_metrics = [col for col in numerical_cols if col in panel.columns]
if key_metrics:
    panel = panel.dropna(subset=key_metrics, how='all')

print(f"Final panel shape: {panel.shape}")

if not panel.empty:
    print("\nSample data from final panel:")
    display_cols = [col for col in [
        col_ticker, col_broker_name, 
        f"{col_shares} {poa_input}", col_target_price, 
        f"{col_market_cap} {poa_input}", f"{col_net_debt} {poa_input}", 
        f"{col_ev} {poa_input}", f"{col_ebitda} {poa_input}", 
        f"{col_ev_ebitda} {poa_input}", f"{col_ebitda_margin} {poa_input}"
    ] if col in panel.columns]
    
    print(panel[display_cols].head())

panel = format_dates(panel)

metrics_to_analyze = [col for col in numerical_cols if col in panel.columns]

if panel.empty:
    print("No data available for analysis")
else:
    print("Creating forecast summary...")
    summary_dfs = create_multi_metric_forecast_summary(
        panel, 
        metrics_to_analyze, 
        output_file="Combined_Forecast_Summary_With_Linking.xlsx"
    )
    
    for ticker, dfs in summary_dfs.items():
        print(f"\n{'='*50}")
        print(f"RESULTS FOR {ticker}")
        print(f"{'='*50}")
        
        print(f"\nForecast Panel Shape: {dfs['Forecast Panel'].shape}")
        print(f"Available Metrics: {list(dfs['Forecast Panel'].columns)}")
        
        key_stats_metrics = [
            f"{col_revenue} {poa_input}", 
            f"{col_ebitda_margin} {poa_input}", 
            f"{col_ev_ebitda} {poa_input}"
        ]
        
        for metric in key_stats_metrics:
            if metric in dfs["Forecast Panel"].columns:
                values = dfs["Forecast Panel"][metric].dropna()
                if not values.empty:
                    try:
                        values = pd.to_numeric(values, errors='coerce').dropna()
                        if not values.empty:
                            p10 = np.percentile(values, 10)
                            median = np.median(values)
                            p90 = np.percentile(values, 90)
                            
                            is_percentage = 'Margin' in metric or 'Dividend Yield' in metric
                            suffix = '%' if is_percentage else ''
                            format_str = "{:.2f}{}" if not is_percentage else "{:.1%}"
                            
                            print(f"\n{metric}:")
                            print(f"  10th percentile: {format_str.format(p10)}")
                            print(f"  Median: {format_str.format(median)}")
                            print(f"  90th percentile: {format_str.format(p90)}")
                    except Exception as e:
                        print(f"Error calculating statistics for {metric}: {e}")
        
        print(f"\nSummary Statistics Table:")
        if not dfs["Summary"].empty:
            print(dfs["Summary"])
        else:
            print("No summary statistics available")

print("\nScript execution completed successfully!")

try:
    rd.close_session()
    print("Refinitiv session closed.")
except Exception as e:
    print(f"Error closing Refinitiv session: {e}")
